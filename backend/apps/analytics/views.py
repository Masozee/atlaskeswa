from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db.models import Count, Sum, Avg, Q
from django.db.models.functions import TruncDate, TruncMonth
from django.utils import timezone
from django.http import HttpResponse
from datetime import timedelta
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from apps.directory.models import Service
from apps.survey.models import Survey, DynamicSurveyResponse
from apps.accounts.models import User
from apps.logs.models import ActivityLog, SystemError
from apps.logs.utils import log_export


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mobile_home(request):
    """
    Get mobile home page data.
    Returns global stats, user-specific stats, user's recent surveys, and kecamatan distribution.
    """
    user = request.user
    from apps.survey.models import QuestionAnswer

    def _survey_stats(qs_old, qs_dynamic):
        old = qs_old.aggregate(
            total=Count('id'),
            draft=Count('id', filter=Q(verification_status=Survey.Status.DRAFT)),
            submitted=Count('id', filter=Q(verification_status=Survey.Status.SUBMITTED)),
            verified=Count('id', filter=Q(verification_status=Survey.Status.VERIFIED)),
            rejected=Count('id', filter=Q(verification_status=Survey.Status.REJECTED)),
        )
        dyn = qs_dynamic.aggregate(
            total=Count('id'),
            draft=Count('id', filter=Q(verification_status=DynamicSurveyResponse.Status.DRAFT)),
            submitted=Count('id', filter=Q(verification_status=DynamicSurveyResponse.Status.SUBMITTED)),
            verified=Count('id', filter=Q(verification_status=DynamicSurveyResponse.Status.VERIFIED)),
            rejected=Count('id', filter=Q(verification_status=DynamicSurveyResponse.Status.REJECTED)),
        )
        return {
            'total':    (old['total']    or 0) + (dyn['total']    or 0),
            'draft':    (old['draft']    or 0) + (dyn['draft']    or 0),
            'submitted':(old['submitted']or 0) + (dyn['submitted']or 0),
            'verified': (old['verified'] or 0) + (dyn['verified'] or 0),
            'rejected': (old['rejected'] or 0) + (dyn['rejected'] or 0),
        }

    global_stats = _survey_stats(Survey.objects.all(), DynamicSurveyResponse.objects.all())
    my_stats = _survey_stats(
        Survey.objects.filter(surveyor=user),
        DynamicSurveyResponse.objects.filter(surveyor=user),
    )

    # Geographic distribution (global, top 5 by Q7 answer)
    kecamatan_distribution = [
        {'kecamatan': item['geographic_unit__name'], 'count': item['count']}
        for item in QuestionAnswer.objects.filter(
            question__code='Q7',
            geographic_unit__isnull=False
        ).exclude(geographic_unit__name='').values('geographic_unit__name').annotate(
            count=Count('id')
        ).order_by('-count')[:5]
    ]

    # User's 5 most recent surveys
    my_dynamic_ids = list(
        DynamicSurveyResponse.objects.filter(surveyor=user).order_by('-created_at')[:5].values_list('id', flat=True)
    )
    my_q7 = {
        ans.response_id: ans.geographic_unit.name if ans.geographic_unit else None
        for ans in QuestionAnswer.objects.filter(
            response_id__in=my_dynamic_ids, question__code='Q7'
        ).select_related('geographic_unit')
    }
    my_old = [
        {
            'id': s.id,
            'service_name': s.service.name if s.service else 'Unknown Service',
            'kecamatan': s.service.kecamatan if s.service else None,
            'city': s.service.city if s.service else None,
            'template_name': None,
            'verification_status': s.verification_status,
            'survey_date': s.survey_date.isoformat() if s.survey_date else None,
            'created_at': s.created_at.isoformat(),
        }
        for s in Survey.objects.filter(surveyor=user).select_related('service').order_by('-created_at')[:5]
    ]
    my_dynamic = [
        {
            'id': s.id,
            'service_name': s.service.name if s.service else 'Unknown Service',
            'kecamatan': s.service.kecamatan if s.service else None,
            'city': s.service.city if s.service else None,
            'template_name': s.template.name if s.template else None,
            'verification_status': s.verification_status,
            'survey_date': s.survey_date.isoformat() if s.survey_date else None,
            'created_at': s.created_at.isoformat(),
            'geographic_unit_name': my_q7.get(s.id),
        }
        for s in DynamicSurveyResponse.objects.filter(surveyor=user).select_related('service', 'template').order_by('-created_at')[:5]
    ]
    recent_surveys = sorted(my_old + my_dynamic, key=lambda x: x['created_at'], reverse=True)[:5]

    return Response({
        'surveys': global_stats,          # kept for backward compat
        'global_surveys': global_stats,
        'my_surveys': my_stats,
        'recent_surveys': recent_surveys,
        'geographic_distribution': kecamatan_distribution,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_stats(request):
    """
    Get comprehensive dashboard statistics
    """

    week_ago = timezone.now() - timedelta(days=7)

    # Consolidate all service stats into a single query
    service_stats = Service.objects.aggregate(
        total=Count('id'),
        verified=Count('id', filter=Q(is_verified=True)),
        active=Count('id', filter=Q(is_active=True)),
        recent=Count('id', filter=Q(created_at__gte=week_ago)),
        total_beds=Sum('bed_capacity'),
        total_staff=Sum('staff_count'),
        total_psychiatrists=Sum('psychiatrist_count'),
        total_psychologists=Sum('psychologist_count'),
        total_nurses=Sum('nurse_count'),
        total_social_workers=Sum('social_worker_count'),
    )

    # Consolidate all survey stats (combine old Survey + DynamicSurveyResponse)
    old_survey_stats = Survey.objects.aggregate(
        total=Count('id'),
        pending=Count('id', filter=Q(verification_status=Survey.Status.SUBMITTED)),
        verified=Count('id', filter=Q(verification_status=Survey.Status.VERIFIED)),
        recent=Count('id', filter=Q(created_at__gte=week_ago)),
    )
    dynamic_survey_stats = DynamicSurveyResponse.objects.aggregate(
        total=Count('id'),
        pending=Count('id', filter=Q(verification_status=DynamicSurveyResponse.Status.SUBMITTED)),
        verified=Count('id', filter=Q(verification_status=DynamicSurveyResponse.Status.VERIFIED)),
        recent=Count('id', filter=Q(created_at__gte=week_ago)),
    )
    survey_stats = {
        'total': old_survey_stats['total'] + dynamic_survey_stats['total'],
        'pending': old_survey_stats['pending'] + dynamic_survey_stats['pending'],
        'verified': old_survey_stats['verified'] + dynamic_survey_stats['verified'],
        'recent': old_survey_stats['recent'] + dynamic_survey_stats['recent'],
    }

    # Consolidate user stats into a single query
    user_stats = User.objects.aggregate(
        total=Count('id'),
        active=Count('id', filter=Q(is_active=True)),
    )

    # Geographic distribution (by kecamatan)
    kecamatan_distribution = Service.objects.filter(
        kecamatan__isnull=False
    ).exclude(kecamatan='').values('kecamatan').annotate(
        count=Count('id')
    ).order_by('-count')

    # MTC distribution - limit in SQL not Python
    mtc_distribution = Service.objects.values(
        'mtc__code', 'mtc__name'
    ).annotate(count=Count('id')).order_by('-count')[:10]

    # Consolidate error stats into a single query
    error_stats = SystemError.objects.aggregate(
        unresolved=Count('id', filter=Q(is_resolved=False)),
        critical=Count('id', filter=Q(severity='CRITICAL', is_resolved=False)),
    )

    # Activity trends (last 14 days) broken down by login and survey submissions
    fourteen_days_ago = timezone.now() - timedelta(days=14)
    daily_activities = ActivityLog.objects.filter(
        timestamp__gte=fourteen_days_ago
    ).annotate(
        day=TruncDate('timestamp')
    ).values('day').annotate(
        count=Count('id'),
        logins=Count('id', filter=Q(action=ActivityLog.Action.LOGIN)),
        submissions=Count('id', filter=Q(action=ActivityLog.Action.SURVEY_SUBMIT)),
    ).order_by('day')

    # Latest 5 surveys with details (combine old + dynamic, sorted by created_at)
    old_surveys = [
        {
            'id': s.id,
            'source': 'survey',
            'service_name': s.service.name if s.service else 'Unknown Service',
            'template_name': None,
            'verification_status': s.verification_status,
            'survey_date': s.survey_date.isoformat() if s.survey_date else None,
            'created_at': s.created_at.isoformat(),
        }
        for s in Survey.objects.select_related('service').order_by('-created_at')[:5]
    ]
    dynamic_surveys = [
        {
            'id': s.id,
            'source': 'dynamic',
            'service_name': s.service.name if s.service else 'Unknown Service',
            'template_name': s.template.name if s.template else None,
            'verification_status': s.verification_status,
            'survey_date': s.survey_date.isoformat() if s.survey_date else None,
            'created_at': s.created_at.isoformat(),
        }
        for s in DynamicSurveyResponse.objects.select_related('service', 'template').order_by('-created_at')[:5]
    ]
    recent_surveys_data = sorted(
        old_surveys + dynamic_surveys,
        key=lambda x: x['created_at'],
        reverse=True
    )[:5]

    return Response({
        'services': {
            'total': service_stats['total'],
            'verified': service_stats['verified'],
            'active': service_stats['active'],
            'recent': service_stats['recent']
        },
        'surveys': {
            'total': survey_stats['total'],
            'pending': survey_stats['pending'],
            'verified': survey_stats['verified'],
            'recent': survey_stats['recent']
        },
        'users': {
            'total': user_stats['total'],
            'active': user_stats['active']
        },
        'capacity': {
            'total_beds': service_stats['total_beds'] or 0,
            'total_staff': service_stats['total_staff'] or 0,
            'psychiatrists': service_stats['total_psychiatrists'] or 0,
            'psychologists': service_stats['total_psychologists'] or 0,
            'nurses': service_stats['total_nurses'] or 0,
            'social_workers': service_stats['total_social_workers'] or 0
        },
        'geographic_distribution': list(kecamatan_distribution),
        'mtc_distribution': list(mtc_distribution),
        'system_health': {
            'unresolved_errors': error_stats['unresolved'],
            'critical_errors': error_stats['critical']
        },
        'activity_trends': list(daily_activities),
        'recent_surveys': recent_surveys_data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def service_analytics(request):
    """
    Get detailed service analytics
    """

    services = Service.objects.all()

    # Service type distribution
    type_distribution = services.values('service_type__name').annotate(
        count=Count('id')
    ).order_by('-count')

    # Insurance coverage
    bpjs_services = services.filter(accepts_bpjs=True).count()
    private_insurance = services.filter(accepts_private_insurance=True).count()

    # Emergency services
    emergency_services = services.filter(accepts_emergency=True).count()
    twentyfour_seven = services.filter(is_24_7=True).count()

    # Average capacity metrics
    avg_metrics = services.aggregate(
        avg_beds=Avg('bed_capacity'),
        avg_staff=Avg('staff_count'),
        avg_psychiatrists=Avg('psychiatrist_count'),
        avg_psychologists=Avg('psychologist_count')
    )

    return Response({
        'type_distribution': list(type_distribution),
        'insurance_coverage': {
            'bpjs': bpjs_services,
            'private': private_insurance
        },
        'emergency_services': {
            'accepts_emergency': emergency_services,
            'twentyfour_seven': twentyfour_seven
        },
        'average_metrics': {
            'beds': round(avg_metrics['avg_beds'] or 0, 2),
            'staff': round(avg_metrics['avg_staff'] or 0, 2),
            'psychiatrists': round(avg_metrics['avg_psychiatrists'] or 0, 2),
            'psychologists': round(avg_metrics['avg_psychologists'] or 0, 2)
        }
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def survey_analytics(request):
    """
    Get detailed survey analytics
    """

    surveys = Survey.objects.all()

    # Status distribution over time (last 6 months)
    six_months_ago = timezone.now() - timedelta(days=180)
    monthly_surveys = surveys.filter(
        created_at__gte=six_months_ago
    ).annotate(
        month=TruncMonth('created_at')
    ).values('month', 'verification_status').annotate(
        count=Count('id')
    ).order_by('month')

    # Average occupancy rate
    from django.db.models import F, FloatField, ExpressionWrapper

    avg_occupancy = surveys.exclude(
        current_bed_capacity=0
    ).annotate(
        occupancy_rate=ExpressionWrapper(
            F('beds_occupied') * 100.0 / F('current_bed_capacity'),
            output_field=FloatField()
        )
    ).aggregate(avg=Avg('occupancy_rate'))

    # Patient demographics
    demographics = surveys.aggregate(
        total_patients=Sum('total_patients_served'),
        male_patients=Sum('patients_male'),
        female_patients=Sum('patients_female'),
        age_0_17=Sum('patients_age_0_17'),
        age_18_64=Sum('patients_age_18_64'),
        age_65_plus=Sum('patients_age_65_plus')
    )

    # Surveyor performance
    surveyor_stats = surveys.values(
        'surveyor__email', 'surveyor__first_name', 'surveyor__last_name'
    ).annotate(
        total_surveys=Count('id'),
        verified=Count('id', filter=Q(verification_status=Survey.Status.VERIFIED)),
        pending=Count('id', filter=Q(verification_status=Survey.Status.SUBMITTED)),
        rejected=Count('id', filter=Q(verification_status=Survey.Status.REJECTED))
    ).order_by('-total_surveys')

    return Response({
        'monthly_trends': list(monthly_surveys),
        'average_occupancy_rate': round(avg_occupancy['avg'] or 0, 2),
        'patient_demographics': demographics,
        'surveyor_performance': list(surveyor_stats)[:10]
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_services_excel(request):
    """
    Export services data to Excel
    """
    # Get query parameters for filtering
    province = request.GET.get('province')
    mtc = request.GET.get('mtc')
    status = request.GET.get('status')

    # Filter services with select_related to avoid N+1 queries
    services = Service.objects.select_related('mtc', 'bsic').all()
    if province and province != 'all':
        services = services.filter(province=province)
    if mtc and mtc != 'all':
        services = services.filter(mtc__code=mtc)
    if status and status != 'all':
        if status == 'VERIFIED':
            services = services.filter(is_verified=True)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Services Data"

    # Define headers
    headers = [
        'Service Name', 'Province', 'City', 'MTC Code', 'BSIC Code',
        'Bed Capacity', 'Staff Count', 'Psychiatrists', 'Psychologists',
        'Nurses', 'Social Workers', 'Verified', 'Active', 'Created Date'
    ]

    # Style for headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write data
    for row_num, service in enumerate(services, 2):
        ws.cell(row=row_num, column=1, value=service.name)
        ws.cell(row=row_num, column=2, value=service.province)
        ws.cell(row=row_num, column=3, value=service.city)
        ws.cell(row=row_num, column=4, value=service.mtc.code if service.mtc else '')
        ws.cell(row=row_num, column=5, value=service.bsic.code if service.bsic else '')
        ws.cell(row=row_num, column=6, value=service.bed_capacity or 0)
        ws.cell(row=row_num, column=7, value=service.staff_count or 0)
        ws.cell(row=row_num, column=8, value=service.psychiatrist_count or 0)
        ws.cell(row=row_num, column=9, value=service.psychologist_count or 0)
        ws.cell(row=row_num, column=10, value=service.nurse_count or 0)
        ws.cell(row=row_num, column=11, value=service.social_worker_count or 0)
        ws.cell(row=row_num, column=12, value='Yes' if service.is_verified else 'No')
        ws.cell(row=row_num, column=13, value='Yes' if service.is_active else 'No')
        ws.cell(row=row_num, column=14, value=service.created_at.strftime('%Y-%m-%d'))

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=yakkum-services-export-{timezone.now().strftime("%Y%m%d")}.xlsx'

    # Log export activity
    log_export(request, 'Service', 'Excel', services.count())

    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_services_csv(request):
    """
    Export services data to CSV
    """
    # Get query parameters for filtering
    province = request.GET.get('province')
    mtc = request.GET.get('mtc')
    status = request.GET.get('status')

    # Filter services with select_related to avoid N+1 queries
    services = Service.objects.select_related('mtc', 'bsic').all()
    if province and province != 'all':
        services = services.filter(province=province)
    if mtc and mtc != 'all':
        services = services.filter(mtc__code=mtc)
    if status and status != 'all':
        if status == 'VERIFIED':
            services = services.filter(is_verified=True)

    # Prepare response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename=yakkum-services-export-{timezone.now().strftime("%Y%m%d")}.csv'

    # Write CSV
    writer = csv.writer(response)

    # Write headers
    writer.writerow([
        'Service Name', 'Province', 'City', 'MTC Code', 'BSIC Code',
        'Bed Capacity', 'Staff Count', 'Psychiatrists', 'Psychologists',
        'Nurses', 'Social Workers', 'Verified', 'Active', 'Created Date'
    ])

    # Write data
    record_count = 0
    for service in services:
        writer.writerow([
            service.name,
            service.province,
            service.city,
            service.mtc.code if service.mtc else '',
            service.bsic.code if service.bsic else '',
            service.bed_capacity or 0,
            service.staff_count or 0,
            service.psychiatrist_count or 0,
            service.psychologist_count or 0,
            service.nurse_count or 0,
            service.social_worker_count or 0,
            'Yes' if service.is_verified else 'No',
            'Yes' if service.is_active else 'No',
            service.created_at.strftime('%Y-%m-%d')
        ])
        record_count += 1

    # Log export activity
    log_export(request, 'Service', 'CSV', record_count)

    return response
