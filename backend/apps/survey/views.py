import csv
import io
import tablib
from decimal import Decimal, InvalidOperation
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Count, Q, Avg, Prefetch
from django.http import HttpResponse, FileResponse
from django.utils import timezone

from .models import (
    Survey, SurveyAttachment, SurveyAuditLog,
    GeographicUnit, SurveyTemplate, QuestionSection, Question,
    QuestionChoice, DynamicSurveyResponse, QuestionAnswer, SurveyPhoto
)
from .serializers import (
    SurveyListSerializer, SurveyDetailSerializer,
    SurveyCreateUpdateSerializer, SurveySubmitSerializer,
    SurveyVerifySerializer, SurveyAttachmentSerializer,
    SurveyAuditLogSerializer, SurveyPhotoSerializer,
    # Dynamic questionnaire serializers
    GeographicUnitSerializer,
    SurveyTemplateListSerializer, SurveyTemplateDetailSerializer,
    QuestionSerializer, QuestionCreateUpdateSerializer,
    QuestionChoiceSerializer, QuestionChoiceCreateUpdateSerializer,
    DynamicSurveyResponseListSerializer, DynamicSurveyResponseDetailSerializer,
    DynamicSurveyResponseCreateSerializer
)
from apps.accounts.permissions import (
    IsAdmin,
    IsSurveyorOrAdmin, IsVerifierOrAdmin, IsSurveyOwnerOrReadOnly,
    CanModifySurveyStatus
)
from apps.accounts.mixins import SurveyorFilterMixin
from apps.logs.utils import (
    log_create, log_update, log_delete,
    log_survey_submit, log_survey_verify, log_survey_reject,
    log_file_upload
)


def _darken(hex_color: str, factor: float = 0.85) -> str:
    """Darken a 6-char hex RGB color by `factor` (0..1) for header backgrounds."""
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    r, g, b = int(r * factor), int(g * factor), int(b * factor)
    return f'{r:02X}{g:02X}{b:02X}'


def _is_effectively_empty_answer(value):
    if value is None:
        return True
    if isinstance(value, str):
        return value == ''
    if isinstance(value, (list, tuple, set)):
        return len(value) == 0
    if isinstance(value, dict):
        return len(value) == 0
    return False


class SurveyViewSet(SurveyorFilterMixin, viewsets.ModelViewSet):
    """
    ViewSet for Survey with verification workflow
    """
    queryset = Survey.objects.select_related(
        'service', 'surveyor', 'assigned_verifier', 'verified_by'
    )
    permission_classes = [IsAuthenticated, IsSurveyOwnerOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]

    filterset_fields = {
        'service': ['exact'],
        'surveyor': ['exact'],
        'verification_status': ['exact'],
        'assigned_verifier': ['exact'],
        'survey_date': ['exact', 'gte', 'lte'],
        'created_at': ['gte', 'lte'],
    }

    search_fields = [
        'service__name', 'surveyor__email', 'surveyor_notes',
        'verifier_notes', 'challenges_faced'
    ]

    ordering_fields = ['survey_date', 'created_at', 'verification_status']
    ordering = ['-survey_date']

    def get_serializer_class(self):
        if self.action == 'list':
            return SurveyListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return SurveyCreateUpdateSerializer
        elif self.action == 'submit':
            return SurveySubmitSerializer
        elif self.action == 'verify':
            return SurveyVerifySerializer
        return SurveyDetailSerializer

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update']:
            return [IsSurveyorOrAdmin(), IsSurveyOwnerOrReadOnly()]
        elif self.action == 'verify':
            return [IsVerifierOrAdmin(), CanModifySurveyStatus()]
        elif self.action == 'submit':
            return [IsSurveyorOrAdmin(), CanModifySurveyStatus()]
        return [IsAuthenticated(), IsSurveyOwnerOrReadOnly()]

    # RBAC Mixin Configuration
    rbac_surveyor_field = 'surveyor'
    rbac_verifier_field = 'assigned_verifier'
    rbac_status_field = 'verification_status'
    rbac_verified_status = 'VERIFIED'
    rbac_submitted_status = 'SUBMITTED'
    rbac_admin_sees_all = True
    rbac_allow_superuser = True

    # get_queryset is now handled by SurveyorFilterMixin
    # The mixin automatically filters based on role:
    # - ADMIN: sees all surveys
    # - SURVEYOR: sees own surveys (surveyor=user)
    # - VERIFIER: sees assigned + submitted surveys
    # - VIEWER: sees only verified surveys

    def perform_create(self, serializer):
        """Set surveyor to current user and log activity"""
        instance = serializer.save(surveyor=self.request.user)
        log_create(self.request, instance, f'Created survey for service: {instance.service.name}')

    def perform_update(self, serializer):
        """Update survey and log activity"""
        instance = serializer.save()
        log_update(self.request, instance, f'Updated survey for service: {instance.service.name}')

    def perform_destroy(self, instance):
        """Delete survey and log activity"""
        log_delete(self.request, instance, f'Deleted survey for service: {instance.service.name}')
        instance.delete()

    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Submit survey for verification"""
        survey = self.get_object()

        # Check if survey can be submitted
        if survey.verification_status != Survey.Status.DRAFT:
            return Response(
                {'detail': 'Only draft surveys can be submitted'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check if user is the surveyor
        if survey.surveyor != request.user and request.user.role != 'ADMIN':
            return Response(
                {'detail': 'Only the surveyor can submit this survey'},
                status=status.HTTP_403_FORBIDDEN
            )

        serializer = SurveySubmitSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Update survey
        survey.verification_status = Survey.Status.SUBMITTED
        survey.submitted_at = timezone.now()

        # Assign verifier if provided
        if 'assigned_verifier' in serializer.validated_data:
            survey.assigned_verifier = serializer.validated_data['assigned_verifier']

        survey.save()

        # Create audit log
        SurveyAuditLog.objects.create(
            survey=survey,
            action=SurveyAuditLog.Action.SUBMITTED,
            user=request.user,
            previous_status=Survey.Status.DRAFT,
            new_status=Survey.Status.SUBMITTED,
            notes='Survey submitted for verification'
        )

        # Log activity
        log_survey_submit(request, survey)

        return Response(SurveyDetailSerializer(survey).data)

    @action(detail=True, methods=['post'])
    def verify(self, request, pk=None):
        """Verify or reject survey"""
        survey = self.get_object()

        # Check if survey can be verified
        if survey.verification_status != Survey.Status.SUBMITTED:
            return Response(
                {'detail': 'Only submitted surveys can be verified'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Strict separation of duties: cannot verify your own submission
        if survey.surveyor == request.user:
            return Response(
                {'detail': 'You cannot verify your own survey submission'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Check if user has permission to verify (VERIFIER or ADMIN)
        if request.user.role not in ['VERIFIER', 'ADMIN']:
            return Response(
                {'detail': 'Only verifiers and admins can verify surveys'},
                status=status.HTTP_403_FORBIDDEN
            )

        serializer = SurveyVerifySerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        action_type = serializer.validated_data['action']
        notes = serializer.validated_data.get('notes', '')

        previous_status = survey.verification_status

        if action_type == 'verify':
            survey.verification_status = Survey.Status.VERIFIED
            survey.verified_by = request.user
            survey.verified_at = timezone.now()
            survey.verifier_notes = notes
            audit_action = SurveyAuditLog.Action.VERIFIED
        else:  # reject
            survey.verification_status = Survey.Status.REJECTED
            survey.rejection_reason = serializer.validated_data.get('rejection_reason', '')
            survey.verifier_notes = notes
            audit_action = SurveyAuditLog.Action.REJECTED

        survey.save()

        # Create audit log
        SurveyAuditLog.objects.create(
            survey=survey,
            action=audit_action,
            user=request.user,
            previous_status=previous_status,
            new_status=survey.verification_status,
            notes=notes
        )

        # Log activity
        if action_type == 'verify':
            log_survey_verify(request, survey)
        else:
            log_survey_reject(request, survey, serializer.validated_data.get('rejection_reason', ''))

        return Response(SurveyDetailSerializer(survey).data)

    @action(detail=True, methods=['get'])
    def attachments(self, request, pk=None):
        """Get all attachments for this survey"""
        survey = self.get_object()
        attachments = SurveyAttachment.objects.filter(survey=survey).select_related('uploaded_by')
        serializer = SurveyAttachmentSerializer(attachments, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def audit_logs(self, request, pk=None):
        """Get audit log for this survey"""
        survey = self.get_object()
        logs = SurveyAuditLog.objects.filter(survey=survey).select_related('user').order_by('-timestamp')
        serializer = SurveyAuditLogSerializer(logs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get survey statistics"""
        queryset = self.get_queryset()

        total_surveys = queryset.count()

        # Status distribution
        status_distribution = queryset.values('verification_status').annotate(
            count=Count('id')
        )

        # Surveys by surveyor
        surveyor_stats = queryset.values(
            'surveyor__email', 'surveyor__first_name', 'surveyor__last_name'
        ).annotate(count=Count('id')).order_by('-count')

        # Average occupancy rate
        avg_occupancy = queryset.exclude(
            current_bed_capacity=0
        ).aggregate(
            avg_occupancy=Avg('beds_occupied') * 100.0 / Avg('current_bed_capacity')
        )['avg_occupancy'] or 0

        # Recent surveys (last 30 days)
        from datetime import timedelta
        thirty_days_ago = timezone.now() - timedelta(days=30)
        recent_surveys = queryset.filter(created_at__gte=thirty_days_ago).count()

        return Response({
            'total_surveys': total_surveys,
            'status_distribution': list(status_distribution),
            'top_surveyors': list(surveyor_stats)[:10],
            'average_occupancy_rate': round(avg_occupancy, 2),
            'recent_surveys': recent_surveys
        })


class SurveyAttachmentViewSet(viewsets.ModelViewSet):
    """
    ViewSet for SurveyAttachment
    """
    queryset = SurveyAttachment.objects.select_related('survey', 'uploaded_by')
    serializer_class = SurveyAttachmentSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['survey', 'attachment_type', 'uploaded_by']
    ordering = ['-uploaded_at']

    def perform_create(self, serializer):
        """Set uploaded_by to current user and log activity"""
        instance = serializer.save(uploaded_by=self.request.user)
        filename = instance.file.name if instance.file else 'unknown'
        log_file_upload(self.request, instance, filename)

    def perform_destroy(self, instance):
        """Delete attachment and log activity"""
        log_delete(self.request, instance, f'Deleted attachment: {instance.file.name}')
        instance.delete()


class SurveyAuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for SurveyAuditLog (read-only)
    """
    queryset = SurveyAuditLog.objects.select_related('survey', 'user')
    serializer_class = SurveyAuditLogSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['survey', 'action', 'user', 'previous_status', 'new_status']
    ordering = ['-timestamp']


# =============================================================================
# DYNAMIC QUESTIONNAIRE VIEWSETS
# =============================================================================

class GeographicUnitViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Geographic Units (full CRUD)
    Provides hierarchical administrative unit data for cascading dropdowns
    """
    queryset = GeographicUnit.objects.select_related(
        'parent__parent__parent'
    )
    serializer_class = GeographicUnitSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    # Note: 'parent' is handled manually in get_queryset due to self-referential FK
    filterset_fields = ['level', 'is_active']
    search_fields = ['name', 'code']
    ordering = ['level', 'name']

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return GeographicUnitCreateUpdateSerializer
        return GeographicUnitSerializer

    def get_permissions(self):
        if self.action in ['import_data']:
            return [IsAdmin()]
        return [IsAuthenticated()]

    def get_queryset(self):
        queryset = super().get_queryset()
        # Handle parent filter explicitly since parent is a self-referential FK
        # and DjangoFilterBackend doesn't handle it correctly
        parent = self.request.query_params.get('parent')
        if parent:
            queryset = queryset.filter(parent_id=parent)
        return queryset

    @action(detail=False, methods=['get'])
    def by_level(self, request):
        """Get units by level with optional parent filter"""
        level = request.query_params.get('level')
        parent_id = request.query_params.get('parent')

        queryset = self.get_queryset()

        if level:
            queryset = queryset.filter(level=level)
        if parent_id:
            queryset = queryset.filter(parent_id=parent_id)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def children(self, request, pk=None):
        """Get children of a geographic unit"""
        unit = self.get_object()
        children = unit.children.filter(is_active=True).order_by('name')
        serializer = self.get_serializer(children, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='export', url_name='export')
    def export_data(self, request):
        """Export geographic units as CSV/XLSX/JSON."""
        fmt = request.query_params.get('file_format', 'csv').lower()
        ids_param = request.query_params.get('ids')
        qs = self.get_queryset().select_related('parent').order_by('level', 'code')
        if ids_param:
            ids = [i for i in ids_param.split(',') if i.strip().isdigit()]
            qs = qs.filter(id__in=ids)

        headers = ['id', 'code', 'name', 'level', 'parent', 'latitude', 'longitude', 'is_active']

        if fmt == 'json':
            import json
            data = []
            for unit in qs:
                data.append({
                    'id': unit.id,
                    'code': unit.code,
                    'name': unit.name,
                    'level': unit.level,
                    'parent': unit.parent.code if unit.parent else '',
                    'latitude': str(unit.latitude) if unit.latitude is not None else '',
                    'longitude': str(unit.longitude) if unit.longitude is not None else '',
                    'is_active': unit.is_active,
                })
            response = HttpResponse(
                json.dumps(data, ensure_ascii=False, indent=2),
                content_type='application/json'
            )
            response['Content-Disposition'] = 'attachment; filename="geographic-units.json"'
            return response

        if fmt == 'xlsx':
            dataset = tablib.Dataset(headers=headers)
            for unit in qs:
                dataset.append([
                    unit.id,
                    unit.code,
                    unit.name,
                    unit.level,
                    unit.parent.code if unit.parent else '',
                    str(unit.latitude) if unit.latitude is not None else '',
                    str(unit.longitude) if unit.longitude is not None else '',
                    unit.is_active,
                ])
            buf = io.BytesIO(dataset.xlsx)
            response = FileResponse(
                buf,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="geographic-units.xlsx"'
            return response

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="geographic-units.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        for unit in qs:
            writer.writerow([
                unit.id,
                unit.code,
                unit.name,
                unit.level,
                unit.parent.code if unit.parent else '',
                str(unit.latitude) if unit.latitude is not None else '',
                str(unit.longitude) if unit.longitude is not None else '',
                unit.is_active,
            ])
        return response

    @action(detail=False, methods=['post'], url_path='import')
    def import_data(self, request):
        """Import geographic units from CSV/XLSX/JSON."""
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        update_existing = request.data.get('update_existing', 'false')
        if isinstance(update_existing, str):
            update_existing = update_existing.lower() in ('true', '1', 'yes')

        filename = file.name.lower()
        try:
            if filename.endswith('.xlsx'):
                dataset = tablib.Dataset().load(file.read(), headers=True, format='xlsx')
                rows = [dict(zip(dataset.headers, row)) for row in dataset]
            elif filename.endswith('.json'):
                import json
                rows = json.loads(file.read().decode('utf-8'))
            else:
                decoded = file.read().decode('utf-8')
                reader = csv.DictReader(io.StringIO(decoded))
                rows = list(reader)
        except Exception as e:
            return Response({'detail': f'Could not read file: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        valid_levels = {choice[0] for choice in GeographicUnit.Level.choices}
        created_count = 0
        updated_count = 0
        errors = []

        for i, raw_row in enumerate(rows, start=2):
            norm = {str(k or '').strip().lower(): '' if v is None else str(v).strip() for k, v in raw_row.items()}
            code = norm.get('code', '')
            name = norm.get('name', '')
            level = norm.get('level', '').upper()
            parent_code = norm.get('parent', '') or norm.get('parent_code', '')
            latitude_raw = norm.get('latitude', '')
            longitude_raw = norm.get('longitude', '')
            is_active_raw = norm.get('is_active', 'true').lower()

            if not code:
                errors.append({'row': i, 'error': 'Missing required field: code'})
                continue
            if not name:
                errors.append({'row': i, 'error': 'Missing required field: name'})
                continue
            if not level:
                errors.append({'row': i, 'error': 'Missing required field: level'})
                continue
            if level not in valid_levels:
                errors.append({'row': i, 'error': f'Invalid level: {level}'})
                continue

            parent = None
            if parent_code:
                parent = GeographicUnit.objects.filter(code=parent_code).first()
                if not parent:
                    errors.append({'row': i, 'error': f'Parent not found with code: {parent_code}'})
                    continue

            latitude = None
            if latitude_raw != '':
                try:
                    latitude = Decimal(latitude_raw)
                except (InvalidOperation, ValueError):
                    errors.append({'row': i, 'error': f'Invalid latitude: {latitude_raw}'})
                    continue

            longitude = None
            if longitude_raw != '':
                try:
                    longitude = Decimal(longitude_raw)
                except (InvalidOperation, ValueError):
                    errors.append({'row': i, 'error': f'Invalid longitude: {longitude_raw}'})
                    continue

            is_active = is_active_raw in ('true', '1', 'yes')
            existing = GeographicUnit.objects.filter(code=code).first()

            if existing:
                if update_existing:
                    existing.name = name
                    existing.level = level
                    existing.parent = parent
                    existing.latitude = latitude
                    existing.longitude = longitude
                    existing.is_active = is_active
                    existing.save()
                    updated_count += 1
                else:
                    errors.append({'row': i, 'error': f'Duplicate code: {code}'})
            else:
                GeographicUnit.objects.create(
                    code=code,
                    name=name,
                    level=level,
                    parent=parent,
                    latitude=latitude,
                    longitude=longitude,
                    is_active=is_active,
                )
                created_count += 1

        return Response({'created': created_count, 'updated': updated_count, 'errors': errors})


class SurveyTemplateViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for Survey Templates (read-only for surveyors)
    Returns templates with their sections and questions
    """
    queryset = SurveyTemplate.objects.filter(is_active=True).annotate(
        total_questions=Count('sections__questions')
    ).prefetch_related(
        'sections__questions__choices'
    )
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['template_type', 'is_active', 'target_mtc']
    search_fields = ['name', 'code', 'description']
    ordering = ['template_type', 'name']

    def get_serializer_class(self):
        if self.action == 'list':
            return SurveyTemplateListSerializer
        return SurveyTemplateDetailSerializer

    @action(detail=True, methods=['get'])
    def sections(self, request, pk=None):
        """Get sections for a template"""
        template = self.get_object()
        sections = template.sections.prefetch_related('questions__choices').order_by('order')
        from .serializers import QuestionSectionSerializer
        serializer = QuestionSectionSerializer(sections, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def questions(self, request, pk=None):
        """Get all questions for a template (flat list)"""
        template = self.get_object()
        questions = Question.objects.filter(
            section__template=template
        ).select_related('section').prefetch_related('choices').order_by(
            'section__order', 'order'
        )
        from .serializers import QuestionSerializer
        serializer = QuestionSerializer(questions, many=True)
        return Response(serializer.data)


class QuestionViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing Questions (ADMIN only)
    Supports CRUD operations for survey template questions
    """
    queryset = Question.objects.select_related('section', 'mtc_code').prefetch_related('choices')
    permission_classes = [IsAuthenticated]
    pagination_class = None
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['section', 'section__template', 'answer_type', 'is_required']
    search_fields = ['code', 'question_text', 'keterangan']
    ordering = ['section__order', 'order']

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return QuestionCreateUpdateSerializer
        return QuestionSerializer

    def get_permissions(self):
        from apps.accounts.permissions import IsAdmin
        if self.action in ['create', 'update', 'partial_update', 'destroy',
                           'bulk_delete', 'import_questions']:
            return [IsAdmin()]
        return [IsAuthenticated()]

    @action(detail=False, methods=['get'], url_path='sections')
    def list_sections(self, request):
        """Return all question sections (lightweight, no nested questions)."""
        sections = QuestionSection.objects.all().order_by('order')
        data = [{'id': s.id, 'code': s.code, 'name': s.name, 'order': s.order, 'template': s.template_id} for s in sections]
        return Response(data)

    @action(detail=False, methods=['post'], url_path='bulk-delete')
    def bulk_delete(self, request):
        """Delete multiple questions by IDs."""
        ids = request.data.get('ids', [])
        if not ids or not isinstance(ids, list):
            return Response({'detail': 'ids must be a non-empty list.'}, status=status.HTTP_400_BAD_REQUEST)
        deleted_count, _ = Question.objects.filter(id__in=ids).delete()
        return Response({'deleted': deleted_count})

    @action(detail=False, methods=['get'], url_path='export', url_name='export')
    def export_data(self, request):
        """Export questions as CSV. Accepts ?ids=1,2,3 and ?file_format=csv|xlsx|json."""
        ids_param = request.query_params.get('ids')
        fmt = request.query_params.get('file_format', 'csv').lower()
        qs = Question.objects.select_related('section', 'mtc_code').order_by('section__order', 'order')
        if ids_param:
            ids = [i for i in ids_param.split(',') if i.strip().isdigit()]
            qs = qs.filter(id__in=ids)

        headers = ['id', 'section_id', 'section_code', 'code', 'question_text',
                   'answer_type', 'is_required', 'order', 'mtc_code', 'desde_ltc_description', 'keterangan']

        if fmt == 'json':
            import json
            data = []
            for q in qs:
                data.append({
                    'id': q.id, 'section_id': q.section_id,
                    'section_code': q.section.code if q.section else '',
                    'code': q.code, 'question_text': q.question_text,
                    'answer_type': q.answer_type, 'is_required': q.is_required,
                    'order': q.order,
                    'mtc_code': q.mtc_code.code if q.mtc_code else '',
                    'desde_ltc_description': q.desde_ltc_description or '',
                    'keterangan': q.keterangan or '',
                })
            response = HttpResponse(json.dumps(data, ensure_ascii=False, indent=2), content_type='application/json')
            response['Content-Disposition'] = 'attachment; filename="questions.json"'
            return response

        if fmt == 'xlsx':
            dataset = tablib.Dataset(headers=headers)
            for q in qs:
                dataset.append([
                    q.id, q.section_id,
                    q.section.code if q.section else '',
                    q.code, q.question_text, q.answer_type, q.is_required,
                    q.order, q.mtc_code.code if q.mtc_code else '',
                    q.desde_ltc_description or '', q.keterangan or '',
                ])
            buf = io.BytesIO(dataset.xlsx)
            response = FileResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="questions.xlsx"'
            return response

        # Default CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="questions.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        for q in qs:
            writer.writerow([
                q.id, q.section_id,
                q.section.code if q.section else '',
                q.code, q.question_text, q.answer_type, q.is_required,
                q.order, q.mtc_code.code if q.mtc_code else '',
                q.desde_ltc_description or '', q.keterangan or '',
            ])
        return response

    @action(detail=False, methods=['post'], url_path='import')
    def import_questions(self, request):
        """Import questions from CSV/XLSX/JSON. Required columns: section_id, code, question_text, answer_type."""
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        update_existing = request.data.get('update_existing', 'false')
        if isinstance(update_existing, str):
            update_existing = update_existing.lower() in ('true', '1', 'yes')

        filename = file.name.lower()
        try:
            if filename.endswith('.xlsx'):
                dataset = tablib.Dataset().load(file.read(), headers=True, format='xlsx')
                rows = [dict(zip(dataset.headers, row)) for row in dataset]
            elif filename.endswith('.json'):
                import json
                rows = json.loads(file.read().decode('utf-8'))
            else:
                decoded = file.read().decode('utf-8')
                reader = csv.DictReader(io.StringIO(decoded))
                rows = list(reader)
        except Exception as e:
            return Response({'detail': f'Could not read file: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        valid_answer_types = {c[0] for c in Question.AnswerType.choices}
        created_count = 0
        updated_count = 0
        errors = []

        for i, raw_row in enumerate(rows, start=2):
            norm = {k.strip().lower(): str(v or '').strip() for k, v in raw_row.items()}
            section_id = norm.get('section_id', '')
            code = norm.get('code', '')
            question_text = norm.get('question_text', '')
            answer_type = norm.get('answer_type', 'TEXT').upper()
            is_required_raw = norm.get('is_required', 'false').lower()
            order_raw = norm.get('order', '0')

            if not code:
                errors.append({'row': i, 'error': 'Missing required field: code'})
                continue
            if not question_text:
                errors.append({'row': i, 'error': 'Missing required field: question_text'})
                continue
            if not section_id:
                errors.append({'row': i, 'error': 'Missing required field: section_id'})
                continue
            if answer_type not in valid_answer_types:
                errors.append({'row': i, 'error': f'Invalid answer_type: {answer_type}'})
                continue

            try:
                section_id_int = int(section_id)
                from .models import QuestionSection
                section = QuestionSection.objects.get(id=section_id_int)
            except (ValueError, QuestionSection.DoesNotExist):
                errors.append({'row': i, 'error': f'Section not found with id: {section_id}'})
                continue

            is_required = is_required_raw in ('true', '1', 'yes')
            try:
                order = int(order_raw)
            except ValueError:
                order = 0

            desde_ltc = norm.get('desde_ltc_description', '')
            keterangan = norm.get('keterangan', '')

            existing = Question.objects.filter(section=section, code=code).first()
            if existing:
                if update_existing:
                    existing.question_text = question_text
                    existing.answer_type = answer_type
                    existing.is_required = is_required
                    existing.order = order
                    existing.desde_ltc_description = desde_ltc
                    existing.keterangan = keterangan
                    existing.save()
                    updated_count += 1
                else:
                    errors.append({'row': i, 'error': f'Duplicate code in section {section_id}: {code}'})
            else:
                Question.objects.create(
                    section=section, code=code, question_text=question_text,
                    answer_type=answer_type, is_required=is_required, order=order,
                    desde_ltc_description=desde_ltc, keterangan=keterangan,
                )
                created_count += 1

        return Response({'created': created_count, 'updated': updated_count, 'errors': errors})


class QuestionSectionViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing Question Sections (ADMIN only)
    Supports CRUD operations for survey template sections
    """
    queryset = QuestionSection.objects.select_related('template').prefetch_related('questions__choices')
    permission_classes = [IsAuthenticated]
    pagination_class = None
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['template']
    search_fields = ['code', 'name']
    ordering = ['template__id', 'order']

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            from .serializers import QuestionSectionCreateUpdateSerializer
            return QuestionSectionCreateUpdateSerializer
        from .serializers import QuestionSectionSerializer
        return QuestionSectionSerializer

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            from apps.accounts.permissions import IsAdmin
            return [IsAdmin()]
        return [IsAuthenticated()]


class QuestionChoiceViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing Question Choices (ADMIN only)
    Supports CRUD operations for question answer choices
    """
    queryset = QuestionChoice.objects.select_related('question', 'mtc_code')
    permission_classes = [IsAuthenticated]
    pagination_class = None
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['question', 'question__section__template']
    search_fields = ['value', 'label', 'keterangan']
    ordering = ['question__section__order', 'question__order', 'order']

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return QuestionChoiceCreateUpdateSerializer
        return QuestionChoiceSerializer

    def get_permissions(self):
        # Only ADMIN can modify choices
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            from apps.accounts.permissions import IsAdmin
            return [IsAdmin()]
        return [IsAuthenticated()]


class DynamicSurveyResponseViewSet(SurveyorFilterMixin, viewsets.ModelViewSet):
    """
    ViewSet for Dynamic Survey Responses with verification workflow
    """
    queryset = DynamicSurveyResponse.objects.select_related(
        'template', 'service', 'surveyor', 'assigned_verifier', 'verified_by',
        'service__mtc', 'service__bsic', 'service__service_type',
    ).prefetch_related(
        'answers__question',
        'answers__selected_choices',
        'answers__geographic_unit__parent__parent__parent',
    )
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]

    filterset_fields = {
        'template': ['exact'],
        'service': ['exact'],
        'surveyor': ['exact'],
        'verification_status': ['exact'],
        'assigned_verifier': ['exact'],
        'deletion_requested': ['exact'],
        'survey_date': ['exact', 'gte', 'lte'],
        'created_at': ['gte', 'lte'],
    }

    search_fields = ['service__name', 'surveyor__email', 'surveyor_notes']
    ordering_fields = ['survey_date', 'created_at', 'verification_status']
    ordering = ['-survey_date']

    # RBAC Mixin Configuration
    rbac_surveyor_field = 'surveyor'
    rbac_verifier_field = 'assigned_verifier'
    rbac_status_field = 'verification_status'
    rbac_verified_status = 'VERIFIED'
    rbac_submitted_status = 'SUBMITTED'
    # ADMIN sees only their own surveys (not all) — aligns with mobile UX requirement
    rbac_admin_sees_all = False
    rbac_allow_superuser = True

    def get_serializer_class(self):
        if self.action == 'list':
            return DynamicSurveyResponseListSerializer
        elif self.action in ['create']:
            return DynamicSurveyResponseCreateSerializer
        return DynamicSurveyResponseDetailSerializer

    def _is_mobile_request(self):
        """Check if request comes from mobile app based on User-Agent."""
        user_agent = self.request.META.get('HTTP_USER_AGENT', '')
        return 'Expo' in user_agent or 'ReactNative' in user_agent

    def _show_all_surveys(self):
        """Check if request should see all surveys (bypass RBAC), e.g. for duplicate checking."""
        return self.request.META.get('HTTP_X_SHOW_ALL', '').lower() in ('true', '1', 'yes')

    def apply_rbac_filter(self, queryset, user):
        """
        Mobile app (Expo/ReactNative): admins see all surveys, other users see only their own.
        Frontend web dashboard: admins see all, other roles use standard RBAC.
        X-Show-All header: bypasses RBAC for duplicate checking across all users.
        """
        # Allow bypass for duplicate checking (mobile app checks which services already have surveys)
        if self._show_all_surveys():
            return queryset

        if self._is_mobile_request():
            # Mobile: admins see all surveys, other roles see only their own
            if user.role == 'ADMIN':
                return queryset
            return queryset.filter(surveyor=user)

        # Frontend web dashboard
        # Admins see all surveys
        if user.role == 'ADMIN':
            return queryset

        # Other roles: use standard RBAC from mixin
        return super().apply_rbac_filter(queryset, user)

    def perform_create(self, serializer):
        """Create survey response with current user as surveyor"""
        instance = serializer.save()
        log_create(self.request, instance, f'Created dynamic survey response for template: {instance.template.name}')

    def perform_update(self, serializer):
        """Update survey response and handle answers if provided"""
        instance = serializer.save()
        gps_update_fields = []
        if 'gps_latitude' in self.request.data:
            instance.latitude = self.request.data.get('gps_latitude')
            gps_update_fields.append('latitude')
        if 'gps_longitude' in self.request.data:
            instance.longitude = self.request.data.get('gps_longitude')
            gps_update_fields.append('longitude')
        if gps_update_fields:
            instance.save(update_fields=gps_update_fields)
        answers_data = self.request.data.get('answers')
        if answers_data and isinstance(answers_data, dict):
            self._update_answers(
                instance,
                answers_data,
                prune_missing=self.request.data.get('prune_missing_answers') is True,
            )
        # Set submitted_at when status transitions to SUBMITTED
        new_status = self.request.data.get('verification_status')
        if new_status == DynamicSurveyResponse.Status.SUBMITTED and not instance.submitted_at:
            instance.submitted_at = timezone.now()
            instance.save(update_fields=['submitted_at'])
        log_update(self.request, instance, f'Updated dynamic survey response for template: {instance.template.name}')

    @action(detail=True, methods=['post'])
    def save_progress(self, request, pk=None):
        """Save progress on a draft survey"""
        response = self.get_object()

        if response.verification_status != DynamicSurveyResponse.Status.DRAFT:
            return Response(
                {'detail': 'Can only save progress on draft surveys'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if response.surveyor != request.user and request.user.role != 'ADMIN':
            return Response(
                {'detail': 'Only the surveyor can update this response'},
                status=status.HTTP_403_FORBIDDEN
            )

        answers_data = request.data.get('answers', {})
        self._update_answers(
            response,
            answers_data,
            prune_missing=request.data.get('prune_missing_answers') is True,
        )

        gps_update_fields = []
        if 'gps_latitude' in request.data:
            response.latitude = request.data.get('gps_latitude')
            gps_update_fields.append('latitude')
        if 'gps_longitude' in request.data:
            response.longitude = request.data.get('gps_longitude')
            gps_update_fields.append('longitude')
        if gps_update_fields:
            response.save(update_fields=gps_update_fields)

        return Response(DynamicSurveyResponseDetailSerializer(response).data)

    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Submit survey for verification"""
        response = self.get_object()

        if response.verification_status != DynamicSurveyResponse.Status.DRAFT:
            return Response(
                {'detail': 'Only draft surveys can be submitted'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if response.surveyor != request.user and request.user.role != 'ADMIN':
            return Response(
                {'detail': 'Only the surveyor can submit this response'},
                status=status.HTTP_403_FORBIDDEN
            )

        response.verification_status = DynamicSurveyResponse.Status.SUBMITTED
        response.submitted_at = timezone.now()
        response.save()

        log_survey_submit(request, response)

        return Response({
            'id': response.id,
            'verification_status': response.verification_status,
            'submitted_at': response.submitted_at,
        })

    @action(detail=True, methods=['post'], parser_classes=[MultiPartParser, FormParser], url_path='upload-answer-file')
    def upload_answer_file(self, request, pk=None):
        """Upload or replace a FILE answer for a specific question/context."""
        response = self.get_object()

        if response.verification_status != DynamicSurveyResponse.Status.DRAFT:
            return Response(
                {'detail': 'Only draft surveys can upload answer files'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if response.surveyor != request.user and request.user.role != 'ADMIN':
            return Response(
                {'detail': 'Only the surveyor can upload answer files for this response'},
                status=status.HTTP_403_FORBIDDEN
            )

        question_code = str(request.data.get('question_code', '')).strip()
        context_key = str(request.data.get('context_key', '')).strip()
        upload = request.FILES.get('file')

        if not question_code:
            return Response({'detail': 'question_code is required'}, status=status.HTTP_400_BAD_REQUEST)

        if not upload:
            return Response({'detail': 'file is required'}, status=status.HTTP_400_BAD_REQUEST)

        question = Question.objects.filter(
            section__template=response.template,
            code=question_code,
        ).first()
        if not question:
            return Response({'detail': 'Question not found in this template'}, status=status.HTTP_404_NOT_FOUND)

        if question.answer_type != Question.AnswerType.FILE:
            return Response({'detail': 'Question is not a FILE answer type'}, status=status.HTTP_400_BAD_REQUEST)

        answer, _ = QuestionAnswer.objects.get_or_create(
            response=response,
            question=question,
            context_key=context_key,
        )

        answer.text_value = ''
        answer.number_value = None
        answer.date_value = None
        answer.time_value = None
        answer.boolean_value = None
        answer.geographic_unit = None
        answer.coverage_level = ''
        answer.table_data = None
        answer.gps_latitude = None
        answer.gps_longitude = None
        answer.other_text = ''
        answer.file = upload
        answer.save()

        answer.selected_choices.clear()

        serializer = QuestionAnswerSerializer(answer, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def verify(self, request, pk=None):
        """Verify or reject survey"""
        response = self.get_object()

        if response.verification_status != DynamicSurveyResponse.Status.SUBMITTED:
            return Response(
                {'detail': 'Only submitted surveys can be verified'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Strict separation of duties: cannot verify your own submission
        if response.surveyor == request.user:
            return Response(
                {'detail': 'You cannot verify your own survey submission'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Check if user has permission to verify (VERIFIER or ADMIN)
        if request.user.role not in ['VERIFIER', 'ADMIN']:
            return Response(
                {'detail': 'Only verifiers and admins can verify surveys'},
                status=status.HTTP_403_FORBIDDEN
            )

        action_type = request.data.get('action')
        notes = request.data.get('notes', '')

        if action_type == 'verify':
            response.verification_status = DynamicSurveyResponse.Status.VERIFIED
            response.verified_by = request.user
            response.verified_at = timezone.now()
            response.verifier_notes = notes
            log_survey_verify(request, response)
        elif action_type == 'reject':
            response.verification_status = DynamicSurveyResponse.Status.REJECTED
            response.rejection_reason = request.data.get('rejection_reason', '')
            response.verifier_notes = notes
            log_survey_reject(request, response, response.rejection_reason)
        else:
            return Response(
                {'detail': 'Invalid action. Use "verify" or "reject"'},
                status=status.HTTP_400_BAD_REQUEST
            )

        response.save()
        return Response(DynamicSurveyResponseDetailSerializer(response).data)

    def get_permissions(self):
        from apps.accounts.permissions import IsAdmin, IsVerifierOrAdmin
        if self.action in ['destroy', 'bulk_delete']:
            return [IsAdmin()]
        if self.action == 'approve_deletion':
            return [IsVerifierOrAdmin()]
        return [IsAuthenticated()]

    @action(detail=False, methods=['post'], url_path='bulk-delete')
    def bulk_delete(self, request):
        """Delete multiple survey responses by IDs."""
        ids = request.data.get('ids', [])
        if not ids or not isinstance(ids, list):
            return Response({'detail': 'Provide a list of IDs'}, status=status.HTTP_400_BAD_REQUEST)
        deleted_count, _ = DynamicSurveyResponse.objects.filter(id__in=ids).delete()
        return Response({'deleted': deleted_count})

    @action(detail=False, methods=['get'], url_path='export', url_name='export')
    def export_data(self, request):
        """Export survey responses as CSV/XLSX.

        Schema (one row per DynamicSurveyResponse):
        - Metadata: ID SURVEY, TGL SURVEY, FASILITAS, ENUM, LONGITUDE, LATITUDE.
        - For every MULTIPLE_CHOICE question Q with choices [v1..vN]:
            * "Q" (or "Q/<ctx>" for DETAIL section): comma-joined selected values
            * "Q/v1", "Q/v2", ... (or "Q/<ctx>/vi"): 1 if selected, else 0
        - For every non-MC question in DETAIL section (the _inline_only_ section):
            * "Q/<ctx>" repeated per upstream choice value, blank when no answer
        - All other questions: single "Q" column.

        Query params:
        - file_format: csv | xlsx (default xlsx)
        - value_format: code | label (default code) — choice cells render as
          choice.value (e.g. "AKUT") or choice.label (full Indonesian text).
        - ids: comma-separated response IDs (optional)
        """
        fmt = request.query_params.get('file_format', 'xlsx').lower()
        value_format = request.query_params.get('value_format', 'code').lower()
        if value_format not in ('code', 'label'):
            value_format = 'code'

        qs = self.filter_queryset(self.get_queryset())
        ids_param = request.query_params.get('ids')
        if ids_param:
            ids = [i for i in ids_param.split(',') if i.strip().isdigit()]
            qs = qs.filter(id__in=ids)

        qs = qs.prefetch_related(
            'answers__question__section',
            'answers__selected_choices',
        )

        headers, plan = self._build_export_schema(qs)
        rows = [self._build_export_row(r, plan, value_format=value_format) for r in qs]

        if fmt == 'csv':
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = 'attachment; filename="surveys.csv"'
            response.write('﻿')  # UTF-8 BOM for Excel
            writer = csv.writer(response)
            writer.writerow(headers)
            writer.writerows(rows)
            return response

        buf = self._build_xlsx_with_colors(headers, plan, rows, qs)
        response = FileResponse(
            buf,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="surveys.xlsx"'
        return response

    @classmethod
    def _build_xlsx_with_colors(cls, headers, plan, rows, responses_qs):
        """Render XLSX via openpyxl with soft per-context column fills.

        Sheets:
          - "Surveys": wide per-question/per-context layout.
          - "Tabel Staff": all STAFF_TABLE entries flattened (ID, KODE,
            JABATAN, LAKI-LAKI, PEREMPUAN).
          - One sheet per KEGIATAN_TABLE / MATRIX_KEGIATAN question (RQJ,
            SRQJ, DQB, DQB1..4, OQB1..2, SDQB*, SOQB*) with flattened rows.

        Per-context columns on Surveys carry pastel fills so each detail
        block (RQA/R1..RQL/R1 vs RQA/R2..RQL/R2) is visually grouped.
        """
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font

        # Soft pastel palette — cycled by index when more contexts than colors.
        palette = [
            'FFF5E6', 'E6F5FF', 'E6FFE6', 'FFE6F0', 'F0E6FF',
            'FFFAE6', 'E6FFFA', 'FAE6FF', 'F5FFE6', 'E6F0FF',
            'FFE6E6', 'E6FFE0', 'FFE0F5', 'E0E6FF', 'FFF0E0',
            'E0FFE6', 'F5E6FF', 'FFFFE0', 'E0FFFF', 'FFE0E0',
        ]

        # Per-column context key extracted from plan payload.
        col_ctx: list[str] = []
        for _label, kind, payload in plan:
            if kind == 'meta':
                col_ctx.append('')
            elif kind in ('value', 'mc_main'):
                col_ctx.append(payload[1] or '')
            elif kind in ('mc_indicator', 'mc_other'):
                col_ctx.append(payload[1] or '')
            else:
                col_ctx.append('')

        # Map ctx → color (first-seen order).
        ctx_to_color: dict[str, str] = {}
        for ctx in col_ctx:
            if ctx and ctx not in ctx_to_color:
                ctx_to_color[ctx] = palette[len(ctx_to_color) % len(palette)]

        wb = Workbook(write_only=False)
        ws = wb.active
        ws.title = 'Surveys'

        # Pre-build PatternFill objects (one per color, reused).
        fill_cache: dict[str, PatternFill] = {
            color: PatternFill(start_color=color, end_color=color, fill_type='solid')
            for color in set(ctx_to_color.values())
        }
        header_fill_cache: dict[str, PatternFill] = {
            color: PatternFill(start_color=_darken(color), end_color=_darken(color), fill_type='solid')
            for color in fill_cache
        }
        header_font = Font(bold=True)

        # Header row.
        ws.append(headers)
        for col_idx, ctx in enumerate(col_ctx, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            color = ctx_to_color.get(ctx)
            if color:
                cell.fill = header_fill_cache[color]

        # Data rows.
        for r_idx, row in enumerate(rows, start=2):
            ws.append(row)
            for c_idx, ctx in enumerate(col_ctx, start=1):
                color = ctx_to_color.get(ctx)
                if color:
                    ws.cell(row=r_idx, column=c_idx).fill = fill_cache[color]

        ws.freeze_panes = 'A2'

        cls._append_table_sheets(wb, responses_qs)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def _append_table_sheets(wb, responses_qs):
        """Add one tab per table-type question, plus aggregated "Tabel Staff".

        STAFF_TABLE rows from every staff-table question go to a single
        "Tabel Staff" sheet keyed by ID SURVEY + KODE (context_key).
        Each KEGIATAN_TABLE / MATRIX_KEGIATAN question gets its own sheet
        named after the question code (e.g. RQJ, DQB, SOQB1).
        """
        from openpyxl.styles import Font
        bold = Font(bold=True)

        # Gather questions of relevant types from involved templates.
        template_ids = list(responses_qs.values_list('template_id', flat=True).distinct())
        if not template_ids:
            return
        table_types = ['STAFF_TABLE', 'KEGIATAN_TABLE', 'OPERATING_HOURS', 'MATRIX_KEGIATAN']
        table_questions = list(
            Question.objects
            .filter(section__template_id__in=template_ids, answer_type__in=table_types)
            .select_related('section')
            .order_by('section__order', 'order')
        )
        if not table_questions:
            return

        # Index answers by question_id → list of (response_id, context_key, table_data)
        relevant_q_ids = [q.id for q in table_questions]
        answers = (
            QuestionAnswer.objects
            .filter(response__in=responses_qs, question_id__in=relevant_q_ids)
            .exclude(table_data__isnull=True)
            .values_list('response_id', 'question_id', 'context_key', 'table_data')
        )
        by_question: dict[int, list] = {}
        for resp_id, qid, ctx, td in answers:
            if not td:
                continue
            by_question.setdefault(qid, []).append((resp_id, ctx, td))

        # ---- Tabel Staff (aggregate all STAFF_TABLE) ----
        staff_rows: list[list] = []
        for q in table_questions:
            if q.answer_type != 'STAFF_TABLE':
                continue
            for resp_id, ctx, td in by_question.get(q.id, []):
                if not isinstance(td, list):
                    continue
                for entry in td:
                    if not isinstance(entry, dict):
                        continue
                    staff_rows.append([
                        resp_id,
                        ctx,
                        entry.get('jabatan', ''),
                        entry.get('laki', entry.get('laki_laki', '')),
                        entry.get('perempuan', ''),
                    ])
        if staff_rows or any(q.answer_type == 'STAFF_TABLE' for q in table_questions):
            ws = wb.create_sheet('Tabel Staff')
            ws.append(['ID SURVEY', 'KODE', 'JABATAN', 'LAKI-LAKI', 'PEREMPUAN'])
            for c in range(1, 6):
                ws.cell(row=1, column=c).font = bold
            for r in staff_rows:
                ws.append(r)
            ws.freeze_panes = 'A2'

        # ---- Per-question sheets (KEGIATAN_TABLE / OPERATING_HOURS / MATRIX_KEGIATAN) ----
        for q in table_questions:
            if q.answer_type == 'STAFF_TABLE':
                continue
            entries = by_question.get(q.id, [])
            if q.answer_type in ('KEGIATAN_TABLE', 'OPERATING_HOURS'):
                hdr = ['ID SURVEY', 'KODE', 'KEGIATAN', 'MULAI', 'SELESAI']
                key_a, key_b = 'start', 'stop'
                name_key = 'kegiatan'
                second_kind = 'time'
            elif q.answer_type == 'MATRIX_KEGIATAN':
                hdr = ['ID SURVEY', 'KODE', 'KEGIATAN', 'HARI']
                key_a = None
                name_key = 'nama_kegiatan'
                second_kind = 'days'
            else:
                continue
            sheet_name = q.code[:31]  # Excel sheet name limit
            ws = wb.create_sheet(sheet_name)
            ws.append(hdr)
            for c in range(1, len(hdr) + 1):
                ws.cell(row=1, column=c).font = bold
            for resp_id, ctx, td in entries:
                if not isinstance(td, list):
                    continue
                for entry in td:
                    if not isinstance(entry, dict):
                        continue
                    name = entry.get(name_key, entry.get('kegiatan', ''))
                    if second_kind == 'time':
                        ws.append([resp_id, ctx, name,
                                   entry.get('start', ''), entry.get('stop', '')])
                    else:  # days
                        days = entry.get('hari', '')
                        if isinstance(days, list):
                            days = ', '.join(days)
                        ws.append([resp_id, ctx, name, days])
            ws.freeze_panes = 'A2'

    @staticmethod
    def _build_export_schema(responses_qs):
        """Return (headers, plan).

        plan is a list of (label, kind, payload). Kinds:
          - 'meta'           : payload = metadata key (str)
          - 'value'          : payload = (code, ctx)               # _format_answer
          - 'mc_main'        : payload = (code, ctx)               # comma-joined values
          - 'mc_indicator'   : payload = (code, ctx, choice_value) # 1 / 0
        """
        meta_labels = ['ID SURVEY', 'TGL SURVEY', 'FASILITAS', 'ENUM',
                       'LONGITUDE', 'LATITUDE']

        template_ids = list(responses_qs.values_list('template_id', flat=True).distinct())
        question_qs = Question.objects.select_related('section').prefetch_related('choices')
        if template_ids:
            question_qs = question_qs.filter(section__template_id__in=template_ids)
        question_qs = question_qs.order_by(
            'section__template_id', 'section__order', 'order'
        )

        questions = list(question_qs)

        # DETAIL = sections marked with show_condition.question_code == "_inline_only_"
        detail_section_ids = set()
        for q in questions:
            sc = q.section.show_condition or {}
            if sc.get('question_code') == '_inline_only_':
                detail_section_ids.add(q.section_id)
        detail_question_codes = {q.code for q in questions
                                 if q.section_id in detail_section_ids}

        # Per-DETAIL-question contexts. A DETAIL question's "family" is the
        # leading uppercase letters before "Q" in its code (RQA→R, SDQB1→SD).
        # A choice value's family is its leading uppercase letters
        # (R12→R, SD3.1.1→SD). Contexts for a DETAIL question = pool values
        # whose family matches.
        import re as _re
        _CODE_FAMILY_RE = _re.compile(r'^([A-Z]+)Q')
        _VALUE_FAMILY_RE = _re.compile(r'^([A-Z]+)')

        def _code_family(code: str) -> str:
            m = _CODE_FAMILY_RE.match(code or '')
            return m.group(1) if m else ''

        def _value_family(val: str) -> str:
            m = _VALUE_FAMILY_RE.match(val or '')
            return m.group(1) if m else ''

        # Build pool: every choice.value targeting any DETAIL question, ordered.
        ctx_pool_seen: set[str] = set()
        ctx_pool_ordered: list[tuple[int, str]] = []
        for q in questions:
            if q.section_id in detail_section_ids:
                continue
            for c in q.choices.all():
                if not c.next_question_code:
                    continue
                if c.next_question_code in detail_question_codes:
                    if c.value not in ctx_pool_seen:
                        ctx_pool_seen.add(c.value)
                        ctx_pool_ordered.append((q.order * 1000 + c.order, c.value))
        ctx_pool_ordered.sort()
        ctx_pool = [v for _, v in ctx_pool_ordered]

        # DETAIL questions in flow order, indexed by family.
        detail_questions_by_family: dict[str, list] = {}
        for q in questions:
            if q.section_id not in detail_section_ids:
                continue
            fam = _code_family(q.code)
            detail_questions_by_family.setdefault(fam, []).append(q)
        # already in section.order, question.order from main queryset

        # Build per-parent-MC detail block plan: parent_question_id -> (contexts, chain)
        # Mirrors mobile flow: when a parent MC's choice points into DETAIL, the
        # detail chain is recursed once per selected value, with context_key=value.
        # Multiple parent MCs in same family each get their OWN inline detail block
        # using only their own contexts.
        parent_detail_blocks: dict[int, tuple[list[str], list]] = {}
        for q in questions:
            if q.section_id in detail_section_ids:
                continue
            choices_into_detail = [
                c for c in sorted(q.choices.all(), key=lambda c: c.order)
                if c.next_question_code in detail_question_codes
            ]
            if not choices_into_detail:
                continue
            # contexts for this parent = its choice values that target DETAIL
            seen_v: set[str] = set()
            contexts: list[str] = []
            families: set[str] = set()
            for c in choices_into_detail:
                if c.value not in seen_v:
                    seen_v.add(c.value)
                    contexts.append(c.value)
                fam = _code_family(c.next_question_code)
                if fam:
                    families.add(fam)
            chain: list = []
            for fam in families:
                chain.extend(detail_questions_by_family.get(fam, []))
            chain.sort(key=lambda dq: (dq.section.order, dq.order))
            parent_detail_blocks[q.id] = (contexts, chain)

        def _emit_question_columns(q, ctx: str = ''):
            """Emit columns for one question (optionally scoped to a context).

            For SINGLE/MULTIPLE_CHOICE questions:
              * main column with comma-joined selected values/labels
              * one indicator column per choice (1/0/blank)
              * extra "_other" column for any choice with has_other_input=True
                — holds QuestionAnswer.other_text when that choice is selected.
            """
            t = q.answer_type
            is_mc = t == Question.AnswerType.MULTIPLE_CHOICE
            is_choice = is_mc or t == Question.AnswerType.SINGLE_CHOICE
            choices = sorted(q.choices.all(), key=lambda c: c.order)
            base = f'{q.code}/{ctx}' if ctx else q.code
            if is_mc:
                plan.append((base, 'mc_main', (q.code, ctx)))
                headers.append(base)
                for c in choices:
                    lab = f'{base}/{c.value}'
                    plan.append((lab, 'mc_indicator', (q.code, ctx, c.value)))
                    headers.append(lab)
                    if c.has_other_input:
                        olab = f'{base}/{c.value}/_other'
                        plan.append((olab, 'mc_other', (q.code, ctx, c.value)))
                        headers.append(olab)
            else:
                plan.append((base, 'value', (q.code, ctx)))
                headers.append(base)
                if is_choice:
                    for c in choices:
                        if c.has_other_input:
                            olab = f'{base}/{c.value}/_other'
                            plan.append((olab, 'mc_other', (q.code, ctx, c.value)))
                            headers.append(olab)

        plan: list = []
        headers: list = []

        for label in meta_labels:
            plan.append((label, 'meta', label))
            headers.append(label)

        seen_codes: set[str] = set()
        emitted_detail: set[tuple[int, str]] = set()  # (detail_q.id, ctx)
        for q in questions:
            if q.section_id in detail_section_ids:
                continue
            if q.code in seen_codes:
                continue
            seen_codes.add(q.code)

            _emit_question_columns(q, ctx='')

            # Inline detail block right after parent MC. Each (detail_q, ctx)
            # pair is emitted at most once — first parent wins. Subsequent
            # parents whose chain overlaps share the existing column.
            block = parent_detail_blocks.get(q.id)
            if not block:
                continue
            contexts, chain = block
            for ctx in contexts:
                for dq in chain:
                    key = (dq.id, ctx)
                    if key in emitted_detail:
                        continue
                    emitted_detail.add(key)
                    _emit_question_columns(dq, ctx=ctx)

        return headers, plan

    def _build_export_row(self, response, plan, value_format: str = 'code'):
        """Build a single export row using the precomputed schema plan.

        value_format='code'  → choice cells use choice.value (e.g. "AKUT").
        value_format='label' → choice cells use choice.label (full text).
        """
        metadata = {
            'ID SURVEY': response.id,
            'TGL SURVEY': response.survey_date.strftime('%Y-%m-%d') if response.survey_date else '',
            'FASILITAS': response.service.name if response.service else '',
            'ENUM': (response.surveyor.get_full_name() or response.surveyor.email) if response.surveyor else '',
            'LONGITUDE': float(response.longitude) if response.longitude is not None else '',
            'LATITUDE': float(response.latitude) if response.latitude is not None else '',
        }

        answer_map: dict[tuple[str, str], QuestionAnswer] = {}
        for ans in response.answers.all():
            answer_map[(ans.question.code, ans.context_key or '')] = ans

        use_label = value_format == 'label'
        row = []
        for _label, kind, payload in plan:
            if kind == 'meta':
                row.append(metadata.get(payload, ''))
            elif kind == 'value':
                ans = answer_map.get(payload)
                row.append(self._format_answer(ans, use_label=use_label) if ans else '')
            elif kind == 'mc_main':
                ans = answer_map.get(payload)
                row.append(','.join(self._mc_selected(ans, use_label=use_label)) if ans else '')
            elif kind == 'mc_indicator':
                code, ctx, value = payload
                ans = answer_map.get((code, ctx))
                if ans is None:
                    row.append('')  # MC unanswered → blank, not 0
                else:
                    row.append(1 if value in self._mc_selected_values(ans) else 0)
            elif kind == 'mc_other':
                code, ctx, value = payload
                ans = answer_map.get((code, ctx))
                if ans and value in self._mc_selected_values(ans):
                    row.append(ans.other_text or '')
                else:
                    row.append('')
            else:
                row.append('')
        return row

    @classmethod
    def _mc_selected(cls, ans, use_label: bool) -> list[str]:
        """Return labels or values for the answer's selected choices."""
        if ans is None:
            return []
        if use_label:
            labels = [c.label for c in ans.selected_choices.all()]
            if labels:
                return labels
        return cls._mc_selected_values(ans)

    @staticmethod
    def _mc_selected_values(ans) -> list[str]:
        """Return list of choice.value strings for an MC/SC answer."""
        if ans is None:
            return []
        values = [c.value for c in ans.selected_choices.all()]
        if values:
            return values
        # Fallback: seeder writes JSON list into text_value for MULTIPLE_CHOICE
        raw = ans.text_value or ''
        if raw.startswith('['):
            try:
                import json as _json
                parsed = _json.loads(raw)
                if isinstance(parsed, list):
                    return [str(v) for v in parsed]
            except (ValueError, TypeError):
                pass
        if raw:
            return [raw]
        return []

    @staticmethod
    def _format_answer(ans, use_label: bool = True):
        """Format a QuestionAnswer to a single cell value.

        use_label=True  → choice answers render as labels (default for plain
        single-question columns to keep them readable).
        use_label=False → choice answers render as choice.value codes.
        """
        q = ans.question
        t = q.answer_type
        if t in ('TEXT', 'TEXTAREA', 'PHONE', 'EMAIL', 'URL'):
            return ans.text_value or ''
        if t in ('NUMBER', 'INTEGER'):
            return float(ans.number_value) if ans.number_value is not None else ''
        if t == 'DATE':
            return ans.date_value.isoformat() if ans.date_value else ''
        if t == 'TIME':
            return ans.time_value.isoformat() if ans.time_value else ''
        if t == 'BOOLEAN':
            if ans.boolean_value is True:
                return 'YA'
            if ans.boolean_value is False:
                return 'TIDAK'
            return ''
        if t in ('SINGLE_CHOICE', 'MULTIPLE_CHOICE'):
            attr = 'label' if use_label else 'value'
            parts = [getattr(c, attr) for c in ans.selected_choices.all()]
            if ans.other_text:
                parts.append(ans.other_text)
            return ', '.join(parts)
        if t.startswith('GEO_'):
            gu = ans.geographic_unit
            if gu:
                getter = getattr(gu, 'get_full_path', None)
                return getter() if callable(getter) else str(gu)
            return ans.text_value or ''
        if t == 'COVERAGE_LEVEL':
            return ans.coverage_level or ''
        return ans.text_value or ''

    @action(detail=True, methods=['post'], url_path='request-deletion')
    def request_deletion(self, request, pk=None):
        """Surveyor requests deletion — must be approved by verifier/admin."""
        response_obj = self.get_object()

        if response_obj.surveyor != request.user and request.user.role != 'ADMIN':
            return Response(
                {'detail': 'Hanya surveyor pemilik yang dapat mengajukan penghapusan'},
                status=status.HTTP_403_FORBIDDEN
            )

        if response_obj.deletion_requested:
            return Response(
                {'detail': 'Permintaan hapus sudah diajukan'},
                status=status.HTTP_400_BAD_REQUEST
            )

        response_obj.deletion_requested = True
        response_obj.deletion_requested_at = timezone.now()
        response_obj.deletion_requested_by = request.user
        response_obj.deletion_reason = request.data.get('reason', '')
        response_obj.save()

        return Response(DynamicSurveyResponseDetailSerializer(response_obj).data)

    @action(detail=True, methods=['post'], url_path='approve-deletion')
    def approve_deletion(self, request, pk=None):
        """Verifier/admin approves or rejects a deletion request."""
        response_obj = self.get_object()

        if not response_obj.deletion_requested:
            return Response(
                {'detail': 'Tidak ada permintaan hapus yang tertunda'},
                status=status.HTTP_400_BAD_REQUEST
            )

        action_type = request.data.get('action')

        if action_type == 'approve':
            response_obj.delete()
            return Response({'detail': 'Survei berhasil dihapus'})
        elif action_type == 'reject':
            response_obj.deletion_requested = False
            response_obj.deletion_requested_at = None
            response_obj.deletion_requested_by = None
            response_obj.deletion_reason = ''
            response_obj.save()
            return Response(DynamicSurveyResponseDetailSerializer(response_obj).data)
        else:
            return Response(
                {'detail': 'Aksi tidak valid. Gunakan "approve" atau "reject"'},
                status=status.HTTP_400_BAD_REQUEST
            )

    def _update_answers(self, response, answers_data, prune_missing=False):
        """Update answers for a response with batched operations"""
        # Separate __other_text keys from regular answers
        other_texts = {}
        clean_answers = {}
        for key, val in answers_data.items():
            if key.endswith('__other_text'):
                other_texts[key.replace('__other_text', '')] = val
            else:
                clean_answers[key] = val
        answers_data = clean_answers

        template = response.template
        questions = Question.objects.filter(section__template=template)
        question_map = {q.code: q for q in questions}

        # Pre-fetch existing answers for this response
        existing_answers = {
            (a.question_id, a.context_key or ''): a
            for a in QuestionAnswer.objects.filter(response=response).select_related('question')
        }

        # Batch fetch all choices for choice-type questions
        choice_question_ids = [
            q.id for q in questions
            if q.answer_type in ['SINGLE_CHOICE', 'MULTIPLE_CHOICE']
        ]
        all_choices = {}
        if choice_question_ids:
            for choice in QuestionChoice.objects.filter(question_id__in=choice_question_ids):
                all_choices.setdefault(choice.question_id, {})[choice.value] = choice

        answers_to_create = []
        answers_to_update = []
        choice_assignments = []  # (answer, question, answer_value)
        incoming_answer_keys = set()

        update_fields = [
            'text_value', 'number_value', 'date_value', 'time_value',
            'boolean_value', 'geographic_unit_id', 'coverage_level',
            'table_data', 'gps_latitude', 'gps_longitude', 'other_text',
        ]

        for raw_code, answer_value in answers_data.items():
            if _is_effectively_empty_answer(answer_value):
                continue

            if '|' in raw_code:
                context_key, question_code = raw_code.split('|', 1)
            else:
                context_key, question_code = '', raw_code

            question = question_map.get(question_code)
            if not question:
                continue
            if question.answer_type == Question.AnswerType.FILE:
                continue
            incoming_answer_keys.add((question.id, context_key))

            # Get existing or create new answer
            answer = existing_answers.get((question.id, context_key))
            is_new = answer is None
            if is_new:
                answer = QuestionAnswer(response=response, question=question, context_key=context_key)
            else:
                answer.context_key = context_key

            # Clear existing values
            answer.text_value = ''
            answer.number_value = None
            answer.date_value = None
            answer.time_value = None
            answer.boolean_value = None
            answer.geographic_unit_id = None
            answer.coverage_level = ''
            answer.table_data = None
            answer.gps_latitude = None
            answer.gps_longitude = None
            answer.other_text = other_texts.get(raw_code, other_texts.get(question_code, ''))

            # Set the appropriate field based on answer type
            if question.answer_type in ['TEXT', 'TEXTAREA', 'PHONE', 'EMAIL', 'URL']:
                answer.text_value = str(answer_value) if answer_value else ''
            elif question.answer_type in ['NUMBER', 'INTEGER']:
                answer.number_value = answer_value
            elif question.answer_type == 'DATE':
                answer.date_value = answer_value
            elif question.answer_type == 'TIME':
                answer.time_value = answer_value
            elif question.answer_type == 'BOOLEAN':
                answer.boolean_value = answer_value
            elif question.answer_type in ['GEO_PROVINSI', 'GEO_KABUPATEN', 'GEO_KECAMATAN']:
                if answer_value:
                    try:
                        answer.geographic_unit_id = int(answer_value)
                    except (TypeError, ValueError):
                        # Mobile sends name string — look up by name
                        level_map = {
                            'GEO_PROVINSI': 'PROVINSI',
                            'GEO_KABUPATEN': 'KABUPATEN_KOTA',
                            'GEO_KECAMATAN': 'KECAMATAN',
                        }
                        geo = GeographicUnit.objects.filter(
                            name__iexact=str(answer_value),
                            level=level_map[question.answer_type]
                        ).first()
                        answer.geographic_unit_id = geo.id if geo else None
                        answer.text_value = str(answer_value)
            elif question.answer_type == 'GEO_DESA':
                if isinstance(answer_value, int) or (isinstance(answer_value, str) and answer_value.isdigit()):
                    answer.geographic_unit_id = int(answer_value)
                else:
                    answer.text_value = str(answer_value) if answer_value else ''
            elif question.answer_type in ['GEO_FULL', 'LOCATION']:
                if answer_value:
                    answer.table_data = answer_value
            elif question.answer_type == 'COVERAGE_LEVEL':
                answer.coverage_level = answer_value
            elif question.answer_type in ['STAFF_TABLE', 'DIAGNOSIS_TABLE', 'REPEATING_TABLE', 'INTERVENTION_MATRIX', 'OPERATING_HOURS', 'KEGIATAN_TABLE', 'MATRIX_KEGIATAN']:
                answer.table_data = answer_value
            elif question.answer_type == 'GPS':
                if isinstance(answer_value, dict):
                    answer.gps_latitude = answer_value.get('latitude')
                    answer.gps_longitude = answer_value.get('longitude')

            if is_new:
                answers_to_create.append(answer)
            else:
                answers_to_update.append(answer)

            # Track choice assignments
            if question.answer_type in ['SINGLE_CHOICE', 'MULTIPLE_CHOICE']:
                choice_assignments.append((answer, question, answer_value, is_new))

        # Bulk create new answers
        if answers_to_create:
            QuestionAnswer.objects.bulk_create(answers_to_create)

        # Bulk update existing answers
        if answers_to_update:
            QuestionAnswer.objects.bulk_update(answers_to_update, update_fields)

        # Set M2M choice relationships using pre-fetched choices
        for answer, question, answer_value, is_new in choice_assignments:
            if is_new:
                # After bulk_create, the answer object has its PK set
                answer.selected_choices.clear()
            else:
                answer.selected_choices.clear()

            question_choices = all_choices.get(question.id, {})
            if isinstance(answer_value, list):
                matched = [question_choices[v] for v in answer_value if v in question_choices]
                if matched:
                    answer.selected_choices.set(matched)
            elif answer_value and answer_value in question_choices:
                answer.selected_choices.add(question_choices[answer_value])

        if prune_missing:
            stale_answer_ids = [
                answer.id
                for key, answer in existing_answers.items()
                if key not in incoming_answer_keys
                and answer.question.answer_type != Question.AnswerType.FILE
            ]
            if stale_answer_ids:
                QuestionAnswer.objects.filter(id__in=stale_answer_ids).delete()


# =============================================================================
# SURVEY PHOTO VIEW SET
# =============================================================================

class SurveyPhotoViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing survey photos.
    Photos can be uploaded by verifier/admin after survey submission.
    """
    queryset = SurveyPhoto.objects.all()
    serializer_class = SurveyPhotoSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['survey']
    ordering_fields = ['uploaded_at']
    http_method_names = ['get', 'post', 'delete', 'head', 'options']

    def get_queryset(self):
        user = self.request.user
        queryset = SurveyPhoto.objects.select_related('uploaded_by', 'survey')

        # Filter by survey if provided
        survey_id = self.request.query_params.get('survey')
        if survey_id:
            queryset = queryset.filter(survey_id=survey_id)

        # Admin/Verifier can see all, others only their own uploads
        if user.role not in ['ADMIN', 'VERIFIER']:
            queryset = queryset.filter(uploaded_by=user)

        return queryset

    def perform_create(self, serializer):
        # Set uploaded_by to current user
        serializer.save(uploaded_by=self.request.user)

    def create(self, request, *args, **kwargs):
        """Handle photo upload"""
        # Check if user is admin/verifier or surveyor (surveyor can upload to their own surveys)
        user = request.user
        survey_id = request.data.get('survey')

        if not survey_id:
            return Response(
                {'detail': 'Survey ID is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get the survey to check ownership
        try:
            survey = DynamicSurveyResponse.objects.get(id=survey_id)
        except DynamicSurveyResponse.DoesNotExist:
            return Response(
                {'detail': 'Survey not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Only admin/verifier or the surveyor can upload photos
        if user.role not in ['ADMIN', 'VERIFIER'] and survey.surveyor != user:
            return Response(
                {'detail': 'You do not have permission to upload photos for this survey'},
                status=status.HTTP_403_FORBIDDEN
            )

        return super().create(request, *args, **kwargs)

    def get_permissions(self):
        from apps.accounts.permissions import IsAdmin, IsVerifierOrAdmin
        if self.action == 'destroy':
            return [IsAdmin()]
        return [IsAuthenticated()]
